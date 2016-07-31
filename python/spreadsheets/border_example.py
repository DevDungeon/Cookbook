from openpyxl import Workbook
from openpyxl.styles import Border, Side

wb = Workbook()
ws = wb.active

ws['A1'] = "255.255.255.255"

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
ws['A1'].border = thin_border

wb.save('cell_border_example.xlsx')