from openpyxl import Workbook
wb = Workbook()

ws = wb.active  # Use default/active sheet
# Or create a new named sheet and set it to active using index
# ws = wb.create_sheet(title="User Information")
# wb.active = 1  # Default sheet was 0, this new sheet is 1

# Direct cell modification
ws['A1'] = "id"
ws['B1'] = "username"
ws['C1'] = "password"

# Add new row at bottom
ws.append(["1337", "NanoDano", "password1"])

# Can use Python datetime objects
import datetime
ws['D1'] = datetime.datetime.now()

# Change sheet tab color
ws.sheet_properties.tabColor = "660000"

wb.save("users.xlsx")  # Write to disk
